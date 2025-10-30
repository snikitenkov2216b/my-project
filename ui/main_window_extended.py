# ui/main_window_extended.py
"""
Расширенное главное окно приложения с поддержкой расчетов поглощения ПГ.
Оптимизированная версия с ленивой загрузкой вкладок.
"""
import logging
import json
from datetime import datetime
from PyQt6.QtWidgets import (
    QMainWindow, QTabWidget, QMenuBar, QMenu,
    QMessageBox, QVBoxLayout, QWidget, QStatusBar,
    QToolBar, QLabel, QProgressBar, QFileDialog
)
from PyQt6.QtCore import Qt, QTimer
from PyQt6.QtGui import QAction, QIcon, QKeySequence

# Импорт конфигурации для динамической загрузки вкладок
from ui.tab_config import (
    EMISSION_TABS_CONFIG,
    ABSORPTION_TABS_CONFIG,
    get_emission_tab_class,
    get_absorption_tab_class
)

# Импорт расширенной фабрики калькуляторов
from calculations.calculator_factory_extended import ExtendedCalculatorFactory

# Импорт вкладки "Своя формула"
from ui.custom_formula_tab import CustomFormulaTab

# Импорт ленивой загрузки вкладок
from ui.lazy_tab_widget import LazyTabWidget


class ExtendedMainWindow(QMainWindow):
    """Главное окно приложения с расширенной функциональностью."""

    def __init__(self):
        super().__init__()
        self.calculator_factory = ExtendedCalculatorFactory()
        self._init_ui()
        self._init_toolbar()
        self._init_statusbar()
        logging.info("Extended GHG Calculator application started")

    def _init_ui(self):
        """Инициализация интерфейса."""
        self.setWindowTitle("Калькулятор парниковых газов - Расширенная версия")
        self.setGeometry(100, 100, 1400, 900)

        # Центральный виджет
        central_widget = QWidget()
        self.setCentralWidget(central_widget)

        # Основной layout
        main_layout = QVBoxLayout(central_widget)

        # Главные вкладки (выбросы vs поглощение)
        self.main_tabs = QTabWidget()

        # Вкладка выбросов
        self.emissions_tabs = QTabWidget()
        self.emissions_tabs.setTabPosition(QTabWidget.TabPosition.West)
        self._init_emission_tabs()

        # Вкладка поглощения
        self.absorption_tabs = QTabWidget()
        self.absorption_tabs.setTabPosition(QTabWidget.TabPosition.West)
        self._init_absorption_tabs()

        # Вкладка "Своя формула"
        self.custom_formula_tab = CustomFormulaTab()

        # Добавляем главные вкладки
        self.main_tabs.addTab(self.emissions_tabs, "📊 Выбросы ПГ")
        self.main_tabs.addTab(self.absorption_tabs, "🌲 Поглощение ПГ")
        self.main_tabs.addTab(self.custom_formula_tab, "🔬 Своя формула")

        main_layout.addWidget(self.main_tabs)

    def _init_emission_tabs(self):
        """
        Инициализация вкладок расчета выбросов.
        Использует ленивую загрузку для оптимизации.
        """
        for category_num, tab_title in EMISSION_TABS_CONFIG:
            # Создаем factory функцию для ленивой загрузки
            def create_tab(cat_num=category_num):
                tab_class = get_emission_tab_class(cat_num)
                if tab_class:
                    calculator = self.calculator_factory.get_calculator(f"Category{cat_num}")
                    if calculator:
                        return tab_class(calculator)
                    else:
                        logging.warning(f"Калькулятор для Category{cat_num} не найден")
                else:
                    logging.warning(f"Класс вкладки для Category{cat_num} не найден")
                return None

            # Создаем ленивую вкладку
            lazy_tab = LazyTabWidget(create_tab, tab_title)
            self.emissions_tabs.addTab(lazy_tab, tab_title)


    def _init_absorption_tabs(self):
        """
        Инициализация вкладок расчета поглощения.
        Использует ленивую загрузку для оптимизации.
        """
        for calc_type, tab_title, module_name, class_name in ABSORPTION_TABS_CONFIG:
            # Создаем factory функцию для ленивой загрузки
            def create_tab(c_type=calc_type, m_name=module_name, c_name=class_name):
                tab_class = get_absorption_tab_class(m_name, c_name)
                if tab_class:
                    calculator = self.calculator_factory.get_absorption_calculator(c_type)
                    if calculator:
                        # Некоторым вкладкам нужен extended_data_service
                        extended_data_service = self.calculator_factory.get_extended_data_service()
                        try:
                            return tab_class(calculator, extended_data_service)
                        except TypeError:
                            return tab_class(calculator)
                    else:
                        logging.warning(f"Калькулятор для {c_type} не найден")
                else:
                    logging.warning(f"Класс вкладки {c_name} из {m_name} не найден")
                return None

            # Создаем ленивую вкладку
            lazy_tab = LazyTabWidget(create_tab, tab_title)
            self.absorption_tabs.addTab(lazy_tab, tab_title)

        # Добавляем сводную вкладку (ленивая загрузка)
        def create_summary_tab():
            summary_class = get_absorption_tab_class("ui.absorption_summary_tab", "AbsorptionSummaryTab")
            if summary_class:
                return summary_class(self.calculator_factory, self.absorption_tabs)
            return None

        lazy_summary = LazyTabWidget(create_summary_tab, "📈 Сводный расчет")
        self.absorption_tabs.addTab(lazy_summary, "📈 Сводный расчет")


    def _init_toolbar(self):
        """Инициализация панели инструментов."""
        toolbar = QToolBar("Главная панель")
        toolbar.setMovable(False)
        self.addToolBar(toolbar)
        # Кнопки быстрого доступа
        new_btn = toolbar.addAction("➕ Новый")
        new_btn.triggered.connect(self._new_calculation)
        open_btn = toolbar.addAction("📂 Открыть")
        open_btn.triggered.connect(self._open_project)
        save_btn = toolbar.addAction("💾 Сохранить")
        save_btn.triggered.connect(self._save_project)
        toolbar.addSeparator()
        balance_btn = toolbar.addAction("⚖️ Баланс")
        balance_btn.triggered.connect(self._show_balance)
        report_btn = toolbar.addAction("📄 Отчет")
        report_btn.triggered.connect(self._export_report)

    def _init_statusbar(self):
        """Инициализация статусной строки."""
        self.statusbar = QStatusBar()
        self.setStatusBar(self.statusbar)
        # Метка состояния
        self.status_label = QLabel("Готово")
        self.statusbar.addWidget(self.status_label)
        # Прогресс-бар для длительных операций
        self.progress_bar = QProgressBar()
        self.progress_bar.setMaximumWidth(200)
        self.progress_bar.setVisible(False)
        self.statusbar.addPermanentWidget(self.progress_bar)
        # Метка с текущим режимом
        self.mode_label = QLabel("Режим: Выбросы ПГ")
        self.statusbar.addPermanentWidget(self.mode_label)
        # Обновление режима при переключении вкладок
        self.main_tabs.currentChanged.connect(self._update_mode_label)

    def _update_mode_label(self, index):
        """Обновление метки режима."""
        if index == 0:
            self.mode_label.setText("Режим: Выбросы ПГ")
        elif index == 1:
            self.mode_label.setText("Режим: Поглощение ПГ")
        elif index == 2:
            self.mode_label.setText("Режим: Своя формула")

    def _new_calculation(self):
        """Начать новый расчет."""
        reply = QMessageBox.question(
            self,
            "Новый расчет",
            "Начать новый расчет? Несохраненные данные будут потеряны.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if reply == QMessageBox.StandardButton.Yes:
            # FIX: Реализована очистка всех полей
            self._clear_all_fields()
            logging.info("Starting new calculation")
            self.status_label.setText("Новый расчет начат")

    def _clear_all_fields(self):
        """Очистка всех полей ввода во всех вкладках."""
        # Очистка вкладок выбросов
        for i in range(self.emissions_tabs.count()):
            tab = self.emissions_tabs.widget(i)
            if hasattr(tab, 'clear_fields'):
                tab.clear_fields()
        # Очистка вкладок поглощения
        for i in range(self.absorption_tabs.count()):
            tab = self.absorption_tabs.widget(i)
            if hasattr(tab, 'clear_fields'):
                tab.clear_fields()
        # Очистка вкладки "Своя формула"
        if hasattr(self, 'custom_formula_tab') and hasattr(self.custom_formula_tab, 'clear_fields'):
            self.custom_formula_tab.clear_fields()
        logging.info("All fields cleared")

    def _open_project(self):
        """Открыть сохраненный проект."""
        try:
            # Диалог выбора файла
            file_path, _ = QFileDialog.getOpenFileName(
                self,
                "Открыть проект",
                "",
                "Файлы проектов (*.ghg *.json);;Все файлы (*.*)"
            )

            if not file_path:
                return  # Пользователь отменил выбор

            self.status_label.setText("Загрузка проекта...")

            # Загрузка данных из файла
            with open(file_path, 'r', encoding='utf-8') as f:
                project_data = json.load(f)

            # Очистка всех полей перед загрузкой
            self._clear_all_fields()

            # Загрузка данных в вкладки выбросов
            if 'emissions' in project_data:
                self._load_emissions_data(project_data['emissions'])

            # Загрузка данных в вкладки поглощения
            if 'absorption' in project_data:
                self._load_absorption_data(project_data['absorption'])

            # Загрузка данных в вкладку "Своя формула"
            if 'custom_formula' in project_data:
                if hasattr(self, 'custom_formula_tab') and hasattr(self.custom_formula_tab, 'set_data'):
                    self.custom_formula_tab.set_data(project_data['custom_formula'])

            self.status_label.setText(f"Проект загружен: {file_path}")
            logging.info(f"Project loaded from {file_path}")

            QMessageBox.information(
                self,
                "Успешно",
                f"Проект успешно загружен из файла:\n{file_path}"
            )

        except json.JSONDecodeError as e:
            logging.error(f"Error parsing project file: {e}")
            QMessageBox.critical(
                self,
                "Ошибка",
                f"Ошибка чтения файла проекта:\n{str(e)}"
            )
            self.status_label.setText("Ошибка загрузки проекта")
        except Exception as e:
            logging.error(f"Error opening project: {e}", exc_info=True)
            QMessageBox.critical(
                self,
                "Ошибка",
                f"Не удалось открыть проект:\n{str(e)}"
            )
            self.status_label.setText("Ошибка загрузки проекта")

    def _save_project(self):
        """Сохранить текущий проект."""
        try:
            # Диалог сохранения файла
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Сохранить проект",
                f"project_{datetime.now().strftime('%Y%m%d_%H%M%S')}.ghg",
                "Файлы проектов (*.ghg *.json);;Все файлы (*.*)"
            )

            if not file_path:
                return  # Пользователь отменил сохранение

            self.status_label.setText("Сохранение проекта...")

            # Сбор данных из всех вкладок
            project_data = {
                'version': '2.0',
                'timestamp': datetime.now().isoformat(),
                'emissions': self._collect_emissions_data(),
                'absorption': self._collect_absorption_data()
            }

            # Сбор данных из вкладки "Своя формула"
            if hasattr(self, 'custom_formula_tab') and hasattr(self.custom_formula_tab, 'get_data'):
                project_data['custom_formula'] = self.custom_formula_tab.get_data()

            # Сохранение в файл
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(project_data, f, ensure_ascii=False, indent=2)

            self.status_label.setText(f"Проект сохранен: {file_path}")
            logging.info(f"Project saved to {file_path}")

            QMessageBox.information(
                self,
                "Успешно",
                f"Проект успешно сохранен в файл:\n{file_path}"
            )

        except Exception as e:
            logging.error(f"Error saving project: {e}", exc_info=True)
            QMessageBox.critical(
                self,
                "Ошибка",
                f"Не удалось сохранить проект:\n{str(e)}"
            )
            self.status_label.setText("Ошибка сохранения проекта")

    def _export_report(self):
        """Экспорт отчета."""
        try:
            # Диалог сохранения отчета
            file_path, selected_filter = QFileDialog.getSaveFileName(
                self,
                "Экспорт отчета",
                f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                "Excel файлы (*.xlsx);;PDF файлы (*.pdf);;HTML файлы (*.html);;Все файлы (*.*)"
            )

            if not file_path:
                return  # Пользователь отменил экспорт

            self.status_label.setText("Создание отчета...")

            # Определяем формат по расширению файла
            if file_path.endswith('.xlsx'):
                self._export_to_excel(file_path)
            elif file_path.endswith('.pdf'):
                self._export_to_pdf(file_path)
            elif file_path.endswith('.html'):
                self._export_to_html(file_path)
            else:
                # По умолчанию экспортируем в Excel
                self._export_to_excel(file_path + '.xlsx')

            self.status_label.setText(f"Отчет экспортирован: {file_path}")
            logging.info(f"Report exported to {file_path}")

            QMessageBox.information(
                self,
                "Успешно",
                f"Отчет успешно экспортирован в файл:\n{file_path}"
            )

        except Exception as e:
            logging.error(f"Error exporting report: {e}", exc_info=True)
            QMessageBox.critical(
                self,
                "Ошибка",
                f"Не удалось экспортировать отчет:\n{str(e)}"
            )
            self.status_label.setText("Ошибка экспорта отчета")

    def _collect_emissions_data(self):
        """Собрать данные из всех вкладок выбросов."""
        emissions_data = {}
        for i in range(self.emissions_tabs.count()):
            tab = self.emissions_tabs.widget(i)
            tab_name = self.emissions_tabs.tabText(i)
            if hasattr(tab, 'get_data'):
                emissions_data[tab_name] = tab.get_data()
        return emissions_data

    def _collect_absorption_data(self):
        """Собрать данные из всех вкладок поглощения."""
        absorption_data = {}
        for i in range(self.absorption_tabs.count()):
            tab = self.absorption_tabs.widget(i)
            tab_name = self.absorption_tabs.tabText(i)
            if hasattr(tab, 'get_data'):
                absorption_data[tab_name] = tab.get_data()
        return absorption_data

    def _load_emissions_data(self, emissions_data):
        """Загрузить данные в вкладки выбросов."""
        logging.info(f"Loading emissions data for {len(emissions_data)} categories")
        for i in range(self.emissions_tabs.count()):
            tab = self.emissions_tabs.widget(i)
            tab_name = self.emissions_tabs.tabText(i)
            if tab_name in emissions_data:
                if hasattr(tab, 'set_data'):
                    tab.set_data(emissions_data[tab_name])
                    logging.info(f"Loaded data for {tab_name}")
                else:
                    logging.warning(f"Tab {tab_name} does not have set_data method")
            else:
                logging.debug(f"No data for {tab_name} in saved file")

    def _load_absorption_data(self, absorption_data):
        """Загрузить данные в вкладки поглощения."""
        for i in range(self.absorption_tabs.count()):
            tab = self.absorption_tabs.widget(i)
            tab_name = self.absorption_tabs.tabText(i)
            if tab_name in absorption_data and hasattr(tab, 'set_data'):
                tab.set_data(absorption_data[tab_name])

    def _export_to_excel(self, file_path):
        """Экспорт отчета в Excel."""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            import re

            # Создаем новую книгу Excel
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # Удаляем стандартный лист

            # Стили
            header_font = Font(bold=True, size=12, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            subheader_font = Font(bold=True, size=11, color="FFFFFF")
            subheader_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")

            total_font = Font(bold=True, size=11)
            total_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Функция для извлечения числа из текста результата
            def extract_number(text):
                if isinstance(text, (int, float)):
                    return float(text)
                if isinstance(text, str):
                    # Ищем число в тексте (поддержка разных форматов)
                    match = re.search(r'[-+]?\d*\.?\d+(?:[eE][-+]?\d+)?', text.replace(',', '.'))
                    if match:
                        try:
                            return float(match.group())
                        except:
                            pass
                return 0.0

            # Создаем лист с общей информацией
            ws_info = wb.create_sheet("Общая информация")
            ws_info['A1'] = "ОТЧЕТ ПО ВЫБРОСАМ И ПОГЛОЩЕНИЮ ПАРНИКОВЫХ ГАЗОВ"
            ws_info['A1'].font = Font(bold=True, size=16, color="1F4E78")
            ws_info.merge_cells('A1:B1')
            ws_info['A3'] = "Дата создания:"
            ws_info['B3'] = datetime.now().strftime('%d.%m.%Y %H:%M:%S')
            ws_info['A4'] = "Версия программы:"
            ws_info['B4'] = "2.0"
            ws_info['A5'] = "Методика:"
            ws_info['B5'] = "Приказ Минприроды РФ от 27.05.2022 N 371"

            # Собираем данные выбросов
            emissions_data = self._collect_emissions_data()
            if emissions_data:
                ws_emissions = wb.create_sheet("Выбросы ПГ")
                row = 1

                # Заголовок (упрощенный - только тип расчета и результат)
                ws_emissions.cell(row, 1, "Тип расчета").font = header_font
                ws_emissions.cell(row, 1).fill = header_fill
                ws_emissions.cell(row, 1).alignment = header_alignment
                ws_emissions.cell(row, 1).border = thin_border

                ws_emissions.cell(row, 2, "Результат (т CO2-экв)").font = header_font
                ws_emissions.cell(row, 2).fill = header_fill
                ws_emissions.cell(row, 2).alignment = header_alignment
                ws_emissions.cell(row, 2).border = thin_border
                row += 1

                total_emissions = 0
                for category, data in emissions_data.items():
                    if isinstance(data, dict):
                        # Извлекаем результат
                        result_text = data.get('result', '')
                        if not result_text or result_text == "Результат появится здесь после расчета":
                            continue  # Пропускаем пустые результаты

                        result_value = extract_number(result_text)
                        if result_value == 0:
                            continue  # Пропускаем нулевые результаты

                        # Название категории
                        ws_emissions.cell(row, 1, category).border = thin_border

                        # Результат
                        ws_emissions.cell(row, 2, result_value).border = thin_border
                        ws_emissions.cell(row, 2).number_format = '0.0000'

                        total_emissions += result_value
                        row += 1

                # Итого выбросов
                if total_emissions > 0:
                    ws_emissions.cell(row, 1, "ИТОГО ВЫБРОСОВ:").font = total_font
                    ws_emissions.cell(row, 1).fill = total_fill
                    ws_emissions.cell(row, 1).border = thin_border

                    ws_emissions.cell(row, 2, total_emissions).font = total_font
                    ws_emissions.cell(row, 2).fill = total_fill
                    ws_emissions.cell(row, 2).border = thin_border
                    ws_emissions.cell(row, 2).number_format = '0.0000'

            # Собираем данные поглощения
            absorption_data = self._collect_absorption_data()
            if absorption_data:
                ws_absorption = wb.create_sheet("Поглощение ПГ")
                row = 1

                # Заголовок (упрощенный - только тип расчета и результат)
                ws_absorption.cell(row, 1, "Тип расчета").font = header_font
                ws_absorption.cell(row, 1).fill = header_fill
                ws_absorption.cell(row, 1).alignment = header_alignment
                ws_absorption.cell(row, 1).border = thin_border

                ws_absorption.cell(row, 2, "Результат (т CO2-экв)").font = header_font
                ws_absorption.cell(row, 2).fill = header_fill
                ws_absorption.cell(row, 2).alignment = header_alignment
                ws_absorption.cell(row, 2).border = thin_border
                row += 1

                total_absorption = 0
                for calc_type, data in absorption_data.items():
                    if isinstance(data, dict):
                        # Извлекаем результат
                        result_text = data.get('result', '')
                        if not result_text or result_text == "Результат появится здесь после расчета":
                            continue  # Пропускаем пустые результаты

                        result_value = extract_number(result_text)
                        if result_value == 0:
                            continue  # Пропускаем нулевые результаты

                        # Название типа расчета
                        ws_absorption.cell(row, 1, calc_type).border = thin_border

                        # Результат
                        ws_absorption.cell(row, 2, result_value).border = thin_border
                        ws_absorption.cell(row, 2).number_format = '0.0000'

                        total_absorption += result_value
                        row += 1

                # Итого поглощения
                if total_absorption > 0:
                    ws_absorption.cell(row, 1, "ИТОГО ПОГЛОЩЕНИЯ:").font = total_font
                    ws_absorption.cell(row, 1).fill = total_fill
                    ws_absorption.cell(row, 1).border = thin_border

                    ws_absorption.cell(row, 2, total_absorption).font = total_font
                    ws_absorption.cell(row, 2).fill = total_fill
                    ws_absorption.cell(row, 2).border = thin_border
                    ws_absorption.cell(row, 2).number_format = '0.0000'

            # Автоматическая ширина столбцов
            for ws in wb.worksheets:
                ws.column_dimensions['A'].width = 40
                ws.column_dimensions['B'].width = 25

            # Сохраняем файл
            wb.save(file_path)
            logging.info(f"Excel report saved to {file_path}")

        except ImportError:
            raise Exception("Библиотека openpyxl не установлена. Установите через: pip install openpyxl")
        except Exception as e:
            logging.error(f"Error creating Excel report: {e}", exc_info=True)
            raise

    def _export_to_pdf(self, file_path):
        """Экспорт отчета в PDF (требует дополнительных библиотек)."""
        # Базовая реализация - можно расширить с помощью reportlab
        raise NotImplementedError("Экспорт в PDF будет реализован в следующей версии. Пожалуйста, используйте формат Excel или HTML.")

    def _export_to_html(self, file_path):
        """Экспорт отчета в HTML."""
        try:
            import re

            # Собираем данные
            emissions_data = self._collect_emissions_data()
            absorption_data = self._collect_absorption_data()

            # Функция для извлечения числа из текста результата
            def extract_number(text):
                if isinstance(text, (int, float)):
                    return float(text)
                if isinstance(text, str):
                    match = re.search(r'[-+]?\d*\.?\d+(?:[eE][-+]?\d+)?', text.replace(',', '.'))
                    if match:
                        try:
                            return float(match.group())
                        except:
                            pass
                return 0.0

            # Формируем HTML
            html_content = f"""
<!DOCTYPE html>
<html lang="ru">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Отчет по выбросам и поглощению ПГ</title>
    <style>
        body {{
            font-family: Arial, sans-serif;
            margin: 20px;
            background-color: #f5f5f5;
        }}
        .container {{
            max-width: 1400px;
            margin: 0 auto;
            background-color: white;
            padding: 30px;
            box-shadow: 0 0 10px rgba(0,0,0,0.1);
        }}
        h1 {{
            color: #1F4E78;
            border-bottom: 3px solid #4472C4;
            padding-bottom: 10px;
        }}
        h2 {{
            color: #34495e;
            margin-top: 30px;
            background-color: #e8f4f8;
            padding: 10px;
            border-left: 5px solid #4472C4;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin: 20px 0;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }}
        th {{
            background-color: #4472C4;
            color: white;
            padding: 12px;
            text-align: left;
            font-weight: bold;
        }}
        td {{
            padding: 10px;
            border: 1px solid #ddd;
            vertical-align: top;
        }}
        tr:nth-child(even) {{
            background-color: #f9f9f9;
        }}
        tr:hover {{
            background-color: #f0f8ff;
        }}
        .total {{
            font-weight: bold;
            background-color: #e8f4f8 !important;
            font-size: 1.1em;
        }}
        .info {{
            background-color: #e8f4f8;
            padding: 15px;
            border-radius: 5px;
            margin: 20px 0;
            border-left: 5px solid #4472C4;
        }}
        .fields {{
            font-size: 0.9em;
            color: #555;
            white-space: pre-wrap;
        }}
        .result {{
            font-weight: bold;
            color: #2c3e50;
            text-align: right;
        }}
    </style>
</head>
<body>
    <div class="container">
        <h1>ОТЧЕТ ПО ВЫБРОСАМ И ПОГЛОЩЕНИЮ ПАРНИКОВЫХ ГАЗОВ</h1>

        <div class="info">
            <p><strong>Дата создания:</strong> {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}</p>
            <p><strong>Версия программы:</strong> 2.0</p>
            <p><strong>Методика:</strong> Приказ Минприроды РФ от 27.05.2022 N 371</p>
        </div>

        <h2>📊 Выбросы ПГ</h2>
        <table>
            <thead>
                <tr>
                    <th style="width: 70%;">Тип расчета</th>
                    <th style="width: 30%;">Результат (т CO2-экв)</th>
                </tr>
            </thead>
            <tbody>
"""
            total_emissions = 0
            for category, data in emissions_data.items():
                if isinstance(data, dict):
                    # Результат
                    result_text = data.get('result', '')
                    if not result_text or result_text == "Результат появится здесь после расчета":
                        continue

                    result_value = extract_number(result_text)
                    if result_value == 0:
                        continue

                    html_content += f"""
                <tr>
                    <td>{category}</td>
                    <td class="result">{result_value:.4f}</td>
                </tr>
"""
                    total_emissions += result_value

            html_content += f"""
                <tr class="total">
                    <td>ИТОГО ВЫБРОСОВ:</td>
                    <td class="result">{total_emissions:.4f}</td>
                </tr>
            </tbody>
        </table>

        <h2>🌲 Поглощение ПГ</h2>
        <table>
            <thead>
                <tr>
                    <th style="width: 70%;">Тип расчета</th>
                    <th style="width: 30%;">Результат (т CO2-экв)</th>
                </tr>
            </thead>
            <tbody>
"""
            total_absorption = 0
            for calc_type, data in absorption_data.items():
                if isinstance(data, dict):
                    # Результат
                    result_text = data.get('result', '')
                    if not result_text or result_text == "Результат появится здесь после расчета":
                        continue

                    result_value = extract_number(result_text)
                    if result_value == 0:
                        continue

                    html_content += f"""
                <tr>
                    <td>{calc_type}</td>
                    <td class="result">{result_value:.4f}</td>
                </tr>
"""
                    total_absorption += result_value

            html_content += f"""
                <tr class="total">
                    <td>ИТОГО ПОГЛОЩЕНИЯ:</td>
                    <td class="result">{total_absorption:.4f}</td>
                </tr>
            </tbody>
        </table>

        <h2>⚖️ Баланс ПГ</h2>
        <div class="info">
            <p><strong>Общие выбросы:</strong> {total_emissions:.4f} т CO2-экв</p>
            <p><strong>Общее поглощение:</strong> {total_absorption:.4f} т CO2-экв</p>
            <p><strong>Чистые выбросы:</strong> {total_emissions - total_absorption:.4f} т CO2-экв</p>
        </div>
    </div>
</body>
</html>
"""
            # Сохраняем HTML файл
            with open(file_path, 'w', encoding='utf-8') as f:
                f.write(html_content)

            logging.info(f"HTML report saved to {file_path}")

        except Exception as e:
            logging.error(f"Error creating HTML report: {e}", exc_info=True)
            raise

    def _calculate_total_emissions(self):
        """
        Рассчитывает суммарные выбросы из всех вкладок.

        Returns:
            tuple: (total_emissions, emissions_by_category, successful_count)
        """
        total_emissions = 0.0
        emissions_by_category = []
        successful_count = 0

        for i in range(self.emissions_tabs.count()):
            tab = self.emissions_tabs.widget(i)
            tab_name = self.emissions_tabs.tabText(i)

            # Получаем результат из вкладки
            if hasattr(tab, 'get_data'):
                data = tab.get_data()
                result_text = data.get('result', '')

                # Извлекаем числовое значение из текста результата
                emission_value = self._extract_number_from_result(result_text)

                if emission_value is not None and emission_value > 0:
                    total_emissions += emission_value
                    emissions_by_category.append({
                        'name': tab_name,
                        'value': emission_value
                    })
                    successful_count += 1
                    logging.debug(f"Emission from {tab_name}: {emission_value:.4f} т CO2-экв")

        logging.info(f"Total emissions calculated: {total_emissions:.4f} т CO2-экв from {successful_count} categories")
        return total_emissions, emissions_by_category, successful_count

    def _calculate_total_absorption(self):
        """
        Рассчитывает суммарное поглощение из всех вкладок.

        Returns:
            tuple: (total_absorption, absorption_by_type, successful_count)
        """
        total_absorption = 0.0
        absorption_by_type = []
        successful_count = 0

        for i in range(self.absorption_tabs.count()):
            tab = self.absorption_tabs.widget(i)
            tab_name = self.absorption_tabs.tabText(i)

            # Пропускаем вкладку "Сводный расчет"
            if "Сводный" in tab_name or "📈" in tab_name:
                continue

            # Получаем результат из вкладки
            if hasattr(tab, 'get_data'):
                data = tab.get_data()
                result_text = data.get('result', '')

                # Извлекаем числовое значение из текста результата
                absorption_value = self._extract_number_from_result(result_text)

                if absorption_value is not None and absorption_value > 0:
                    total_absorption += absorption_value
                    absorption_by_type.append({
                        'name': tab_name,
                        'value': absorption_value
                    })
                    successful_count += 1
                    logging.debug(f"Absorption from {tab_name}: {absorption_value:.4f} т CO2-экв")

        logging.info(f"Total absorption calculated: {total_absorption:.4f} т CO2-экв from {successful_count} types")
        return total_absorption, absorption_by_type, successful_count

    def _extract_number_from_result(self, result_text):
        """
        Извлекает числовое значение из текста результата.

        Args:
            result_text: Строка с результатом (например, "Результат: 1234.56 тонн CO2")

        Returns:
            float или None: Извлеченное число или None, если не найдено
        """
        if not result_text or not isinstance(result_text, str):
            return None

        import re
        # Ищем числа с плавающей точкой в тексте
        # Поддерживаем форматы: 1234.56, 1,234.56, 1234,56
        patterns = [
            r'(\d+(?:[.,]\d+)?)\s*(?:тонн|т|т CO2|тCO2|кг|kg)',
            r'[Рр]езультат[:\s]+(\d+(?:[.,]\d+)?)',
            r'(\d+(?:[.,]\d+))'
        ]

        for pattern in patterns:
            match = re.search(pattern, result_text)
            if match:
                try:
                    number_str = match.group(1).replace(',', '.')
                    return float(number_str)
                except:
                    continue

        return None

    def _show_balance(self):
        """Показать баланс выбросов и поглощения."""
        try:
            # Рассчитываем суммарные значения
            total_emissions, emissions_by_category, em_count = self._calculate_total_emissions()
            total_absorption, absorption_by_type, ab_count = self._calculate_total_absorption()
            net_emissions = total_emissions - total_absorption

            # Создаем красивое диалоговое окно
            from PyQt6.QtWidgets import QDialog, QVBoxLayout, QHBoxLayout, QLabel, QTextEdit, QPushButton, QGroupBox
            from PyQt6.QtCore import Qt

            dialog = QDialog(self)
            dialog.setWindowTitle("⚖️ Баланс парниковых газов")
            dialog.setMinimumWidth(700)
            dialog.setMinimumHeight(600)

            layout = QVBoxLayout(dialog)

            # Заголовок
            title_label = QLabel("<h2>⚖️ Баланс парниковых газов (ПГ)</h2>")
            title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
            layout.addWidget(title_label)

            # Группа: Суммарные значения
            summary_group = QGroupBox("📊 Сводные показатели")
            summary_layout = QVBoxLayout(summary_group)

            emissions_label = QLabel(f"<b>Общие выбросы:</b> <span style='color: #d32f2f; font-size: 18px;'>{total_emissions:.4f}</span> т CO2-экв<br>"
                                    f"<small>(из {em_count} категорий)</small>")
            emissions_label.setTextFormat(Qt.TextFormat.RichText)
            summary_layout.addWidget(emissions_label)

            absorption_label = QLabel(f"<b>Общее поглощение:</b> <span style='color: #388e3c; font-size: 18px;'>{total_absorption:.4f}</span> т CO2-экв<br>"
                                      f"<small>(из {ab_count} типов)</small>")
            absorption_label.setTextFormat(Qt.TextFormat.RichText)
            summary_layout.addWidget(absorption_label)

            # Линия разделитель
            line_label = QLabel("<hr>")
            summary_layout.addWidget(line_label)

            # Чистые выбросы
            net_color = "#d32f2f" if net_emissions > 0 else "#388e3c"
            net_text = "Чистые выбросы" if net_emissions > 0 else "Чистое поглощение"
            net_label = QLabel(f"<b>{net_text}:</b> <span style='color: {net_color}; font-size: 20px; font-weight: bold;'>{abs(net_emissions):.4f}</span> т CO2-экв")
            net_label.setTextFormat(Qt.TextFormat.RichText)
            summary_layout.addWidget(net_label)

            # Процент компенсации
            if total_emissions > 0:
                compensation_percent = (total_absorption / total_emissions) * 100
                compensation_label = QLabel(f"<small><i>Процент компенсации выбросов: {compensation_percent:.2f}%</i></small>")
                compensation_label.setTextFormat(Qt.TextFormat.RichText)
                summary_layout.addWidget(compensation_label)

            layout.addWidget(summary_group)

            # Группа: Детализация по категориям
            details_group = QGroupBox("📋 Детализация")
            details_layout = QVBoxLayout(details_group)

            details_text = QTextEdit()
            details_text.setReadOnly(True)
            details_content = ""

            if emissions_by_category:
                details_content += "<h3 style='color: #d32f2f;'>🔴 Выбросы по категориям:</h3><ul>"
                for item in sorted(emissions_by_category, key=lambda x: x['value'], reverse=True):
                    details_content += f"<li><b>{item['name']}:</b> {item['value']:.4f} т CO2-экв</li>"
                details_content += "</ul><br>"
            else:
                details_content += "<p><i>Нет данных по выбросам. Выполните расчеты в категориях выбросов.</i></p><br>"

            if absorption_by_type:
                details_content += "<h3 style='color: #388e3c;'>🟢 Поглощение по типам:</h3><ul>"
                for item in sorted(absorption_by_type, key=lambda x: x['value'], reverse=True):
                    details_content += f"<li><b>{item['name']}:</b> {item['value']:.4f} т CO2-экв</li>"
                details_content += "</ul>"
            else:
                details_content += "<p><i>Нет данных по поглощению. Выполните расчеты в разделе поглощения.</i></p>"

            details_text.setHtml(details_content)
            details_layout.addWidget(details_text)

            layout.addWidget(details_group)

            # Кнопки
            button_layout = QHBoxLayout()

            export_btn = QPushButton("📤 Экспортировать отчет")
            export_btn.clicked.connect(lambda: (dialog.accept(), self._export_report()))
            export_btn.setStyleSheet("QPushButton { padding: 8px; }")

            close_btn = QPushButton("✖ Закрыть")
            close_btn.clicked.connect(dialog.accept)
            close_btn.setStyleSheet("QPushButton { padding: 8px; }")

            button_layout.addWidget(export_btn)
            button_layout.addWidget(close_btn)
            layout.addLayout(button_layout)

            # Показываем диалог
            dialog.exec()

            logging.info("Balance dialog shown successfully")

        except Exception as e:
            logging.error(f"Error showing balance: {e}", exc_info=True)
            QMessageBox.critical(
                self,
                "Ошибка",
                f"Не удалось рассчитать баланс:\n{str(e)}"
            )

    def _show_comparison(self):
        """Показать сравнение периодов."""
        QMessageBox.information(self, "В разработке", "Функция сравнения периодов в разработке")

    def _show_settings(self):
        """Показать настройки."""
        QMessageBox.information(self, "В разработке", "Настройки программы в разработке")

    def _show_docs(self):
        """Показать документацию."""
        QMessageBox.information(
            self,
            "Документация",
            "Калькулятор парниковых газов v2.0\n\n"
            "Программа реализует методики расчета:\n"
            "• Выбросов ПГ (25 категорий)\n"
            "• Поглощения ПГ (формулы 1-100)\n\n"
            "Согласно Приказу Минприроды РФ от 27.05.2022 N 371"
        )

    def _show_methodology(self):
        """Показать методологию расчетов."""
        QMessageBox.information(
            self,
            "Методология",
            "Методология расчетов основана на:\n\n"
            "1. Приказе Минприроды РФ от 27.05.2022 N 371\n"
            "2. Руководящих принципах МГЭИК 2006 г.\n"
            "3. Национальных коэффициентах выбросов\n\n"
            "Все формулы и коэффициенты соответствуют\n"
            "официальной методике РФ"
        )

    def _show_about(self):
        """Показать информацию о программе."""
        QMessageBox.about(
            self,
            "О программе",
            "Калькулятор парниковых газов v2.0\n\n"
            "Программа для расчета выбросов и поглощения\n"
            "парниковых газов согласно методике\n"
            "Минприроды России\n\n"
            "© 2024 GHG Calculator Team"
        )